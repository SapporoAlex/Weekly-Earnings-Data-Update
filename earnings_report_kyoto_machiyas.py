from bs4 import BeautifulSoup
from datetime import datetime
from email.message import EmailMessage
import openpyxl as xl
from openpyxl.styles import PatternFill, Font, Border, Side
import os
import pandas as pd
from playwright.sync_api import sync_playwright
import smtplib
import ssl


# date details
current_datetime = datetime.now()
current_date = current_datetime.date()
current_year = current_datetime.year
current_month = current_datetime.month
current_day = current_datetime.day
date_object = datetime(current_year, current_month, current_day)
current_week_of_year = date_object.isocalendar()[1]
next_year = current_year + 1
previous_year = current_year - 1
previous_week_of_year = current_week_of_year - 1

# login details
zeniyacho_contract_code = os.environ.get('ZENIYACHO_CONTRACT_CODE')
zeniyacho_account_name = os.environ.get('ZENIYACHO_ACCOUNT_NAME')
zeniyacho_password = os.environ.get('ZENIYACHO_PASSWORD')
# URL for Zeniyacho DB
zeniyacho_neppan_site_url = '####'

fukune_contract_code = os.environ.get('FUKUNE_CONTRACT_CODE')
fukune_account_name = os.environ.get('FUKUNE_ACCOUNT_NAME')
fukune_password = os.environ.get('FUKUNE_PASSWORD')
# URL for Fukune DB
fukune_neppan_site_url = '####'

login_code_box = "#clientCode"
login_id_box = "#loginId"
login_password_box = "#password"
login_button = "#LoginBtn"
bunseki_button = "#menu > li:nth-child(6)"
bunseki_button_dropdown = "#menu > li:nth-child(6) > ul > li:nth-child(4) > a"
expand_button = "#hideKensakuDispButton"
date_selection_button = "#cmbFromYM"
uriage_button = "#kensakuDisp2 > tbody > tr:nth-child(3) > td > table > tbody > tr > td:nth-child(3) > label"
bunseki_wo_jikou_suru_button = "#SearchBtn"
table = "#kensakuDisp4 > tbody > tr:nth-child(7) > td > table > tbody"
fukune_values_for_empty_column = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
zeniyacho_values_for_empty_column = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]


def main():

    def arrowdown_twelve_times():
        global page
        for i in range(12):
            page.keyboard.press('ArrowDown')
        page.keyboard.press('Enter')

    def get_tables(hotel_name, contract_code, account_name, password, url, sheet_week):
        global current_column_15_sum, next_column_15_sum
        with sync_playwright() as p:
            page_url = url
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(page_url)
            page.is_visible(login_button)
            page.fill(login_code_box, contract_code)
            page.fill(login_id_box, account_name)
            page.fill(login_password_box, password)
            page.click(login_button)
            page.is_visible(bunseki_button)
            page.wait_for_timeout(1000)
            page.click(bunseki_button)
            page.click(bunseki_button_dropdown)
            page.wait_for_timeout(1000)
            page.is_visible(expand_button)
            page.click(expand_button)
            page.is_visible(date_selection_button)
            page.click(date_selection_button)
            page.wait_for_timeout(1000)
            if current_month == 1:
                page.keyboard.press('Enter')
            if current_month == 2:
                page.keyboard.press('ArrowUp')
                page.keyboard.press('Enter')
            if current_month == 3:
                for i in range(2):
                    page.keyboard.press('ArrowUp')
                page.keyboard.press('Enter')
            if current_month == 4:
                for i in range(3):
                    page.keyboard.press('ArrowUp')
                page.keyboard.press('Enter')
            if current_month == 5:
                for i in range(4):
                    page.keyboard.press('ArrowUp')
                page.keyboard.press('Enter')
            if current_month == 6:
                for i in range(5):
                    page.keyboard.press('ArrowUp')
                page.keyboard.press('Enter')
            if current_month == 7:
                for i in range(6):
                    page.keyboard.press('ArrowUp')
                page.keyboard.press('Enter')
            if current_month == 8:
                for i in range(7):
                    page.keyboard.press('ArrowUp')
                page.keyboard.press('Enter')
            if current_month == 9:
                for i in range(8):
                    page.keyboard.press('ArrowUp')
                page.keyboard.press('Enter')
            if current_month == 10:
                for i in range(9):
                    page.keyboard.press('ArrowUp')
                page.keyboard.press('Enter')
            if current_month == 11:
                for i in range(10):
                    page.keyboard.press('ArrowUp')
                page.keyboard.press('Enter')
            if current_month == 12:
                for i in range(11):
                    page.keyboard.press('ArrowUp')
                page.keyboard.press('Enter')
            page.click(uriage_button)
            page.wait_for_timeout(5000)
            page.click(bunseki_wo_jikou_suru_button)
            page.wait_for_timeout(5000)
            html = page.inner_html(table)
            soup = BeautifulSoup(html, 'html.parser')

            table_data = []
            for row in soup.select('tr'):
                row_data = [cell.text.strip() for cell in row.find_all(['th', 'td'])]
                table_data.append(row_data)

            current_df = pd.DataFrame(table_data[1:], columns=table_data[0])
            current_df.iloc[:, 1:] = current_df.iloc[:, 1:].apply(lambda x: x.str.replace(',', ''), axis=1)
            current_df.iloc[:, 1:] = current_df.iloc[:, 1:].apply(pd.to_numeric, errors='coerce')
            current_df["Row_Sum"] = current_df.iloc[:, 1:].sum(axis=1)
            current_df.fillna(0, inplace=True)
            if hotel_name == 'zeniyacho':
                column_13_sum_except_rows_1_and_9 = current_df.loc[(current_df.index != 0)
                                                                   & (current_df.index != 8), 'Row_Sum'].sum()
                current_df.loc[8, "Row_Sum - 10%"] = current_df.iloc[8, 1:-1].sum() * 0.9
                current_total = column_13_sum_except_rows_1_and_9 + current_df.loc[8, "Row_Sum - 10%"]

                row_8_10 = []
                row_8 = current_df.iloc[8, 1:13]
                for i in row_8:
                    i = i * 0.1
                    i = int(i)
                    row_8_10.append(i)
                row_1 = current_df.iloc[0, 1:13]
                result = []
                for a, b in zip(row_1, row_8_10):
                    result.append(a - b)
                current_df.iloc[0, 1:13] = result
                
                
            else:
                column_13_sum_except_rows_1_and_9 = current_df.loc[(current_df.index != 0)
                                                                   & (current_df.index != 7), 'Row_Sum'].sum()
                current_df.loc[7, "Row_Sum - 10%"] = current_df.iloc[7, 1:-1].sum() * 0.9
                current_total = column_13_sum_except_rows_1_and_9 + current_df.loc[7, "Row_Sum - 10%"]
                
                row_7_10 = []
                row_7 = current_df.iloc[7, 1:13]
                for i in row_7:
                    i = i * 0.1
                    i = int(i)
                    row_7_10.append(i)
                row_1 = current_df.iloc[0, 1:13]
                result = []
                for a, b in zip(row_1, row_7_10):
                    result.append(a - b)
                current_df.iloc[0, 1:13] = result

            current_column_15_sum = current_total
            if not page.is_visible(date_selection_button):
                page.click(expand_button)
                page.wait_for_timeout(5000)
                page.click(date_selection_button)
                print('needed to press expand again')
            else:
                page.click(date_selection_button)
                page.wait_for_timeout(5000)
                print('without issue')

            if current_month == 1:
                for i in range(6):
                    page.keyboard.press('ArrowDown')
                page.keyboard.press('Enter')
            if current_month == 2:
                for i in range(7):
                    page.keyboard.press('ArrowDown')
                page.keyboard.press('Enter')
            if current_month == 3:
                for i in range(8):
                    page.keyboard.press('ArrowDown')
                page.keyboard.press('Enter')
            if current_month == 4:
                for i in range(9):
                    page.keyboard.press('ArrowDown')
                page.keyboard.press('Enter')
            if current_month == 5:
                for i in range(10):
                    page.keyboard.press('ArrowDown')
                page.keyboard.press('Enter')
            if current_month == 6:
                for i in range(11):
                    page.keyboard.press('ArrowDown')
                page.keyboard.press('Enter')
            if current_month == 7:
                arrowdown_twelve_times()
            if current_month == 8:
                arrowdown_twelve_times()
            if current_month == 9:
                arrowdown_twelve_times()
            if current_month == 10:
                arrowdown_twelve_times()
            if current_month == 11:
                arrowdown_twelve_times()
            if current_month == 12:
                arrowdown_twelve_times()
            page.click(uriage_button)
            page.wait_for_timeout(5000)
            page.click(bunseki_wo_jikou_suru_button)
            page.wait_for_timeout(5000)
            html = page.inner_html(table)
            soup = BeautifulSoup(html, 'html.parser')

            next_table_data = []
            for row in soup.select('tr'):
                row_data = [cell.text.strip() for cell in row.find_all(['th', 'td'])]
                next_table_data.append(row_data)

            next_df = pd.DataFrame(next_table_data[1:], columns=next_table_data[0])
            if current_month == 1:
                next_df = next_df.drop(columns=['7月', '8月', '9月', '10月', '11月', '12月'])
                if hotel_name == 'zeniyacho':
                    next_df['7月'] = zeniyacho_values_for_empty_column
                    next_df["8月"] = zeniyacho_values_for_empty_column
                    next_df["9月"] = zeniyacho_values_for_empty_column
                    next_df['10月'] = zeniyacho_values_for_empty_column
                    next_df["11月"] = zeniyacho_values_for_empty_column
                    next_df["12月"] = zeniyacho_values_for_empty_column
                else:
                    next_df['7月'] = fukune_values_for_empty_column
                    next_df["8月"] = fukune_values_for_empty_column
                    next_df["9月"] = fukune_values_for_empty_column
                    next_df['10月'] = fukune_values_for_empty_column
                    next_df["11月"] = fukune_values_for_empty_column
                    next_df["12月"] = fukune_values_for_empty_column
            if current_month == 2:
                next_df = next_df.drop(columns=['8月', '9月', '10月', '11月', '12月'])
                if hotel_name == 'zeniyacho':
                    next_df["8月"] = zeniyacho_values_for_empty_column
                    next_df["9月"] = zeniyacho_values_for_empty_column
                    next_df['10月'] = zeniyacho_values_for_empty_column
                    next_df["11月"] = zeniyacho_values_for_empty_column
                    next_df["12月"] = zeniyacho_values_for_empty_column
                else:
                    next_df["8月"] = fukune_values_for_empty_column
                    next_df["9月"] = fukune_values_for_empty_column
                    next_df['10月'] = fukune_values_for_empty_column
                    next_df["11月"] = fukune_values_for_empty_column
                    next_df["12月"] = fukune_values_for_empty_column
            if current_month == 3:
                next_df = next_df.drop(columns=['9月', '10月', '11月', '12月'])
                if hotel_name == 'zeniyacho':
                    next_df["9月"] = zeniyacho_values_for_empty_column
                    next_df['10月'] = zeniyacho_values_for_empty_column
                    next_df["11月"] = zeniyacho_values_for_empty_column
                    next_df["12月"] = zeniyacho_values_for_empty_column
                else:
                    next_df["9月"] = fukune_values_for_empty_column
                    next_df['10月'] = fukune_values_for_empty_column
                    next_df["11月"] = fukune_values_for_empty_column
                    next_df["12月"] = fukune_values_for_empty_column
            if current_month == 4:
                next_df = next_df.drop(columns=['10月', '11月', '12月'])
                if hotel_name == 'zeniyacho':
                    next_df['10月'] = zeniyacho_values_for_empty_column
                    next_df["11月"] = zeniyacho_values_for_empty_column
                    next_df["12月"] = zeniyacho_values_for_empty_column
                else:
                    next_df['10月'] = fukune_values_for_empty_column
                    next_df["11月"] = fukune_values_for_empty_column
                    next_df["12月"] = fukune_values_for_empty_column
            if current_month == 5:
                next_df = next_df.drop(columns=['11月', '12月'])
                if hotel_name == 'zeniyacho':
                    next_df["11月"] = zeniyacho_values_for_empty_column
                    next_df["12月"] = zeniyacho_values_for_empty_column
                else:
                    next_df["11月"] = fukune_values_for_empty_column
                    next_df["12月"] = fukune_values_for_empty_column
            if current_month == 6:
                next_df = next_df.drop(columns=['12月'])
                if hotel_name == 'zeniyacho':
                    next_df["12月"] = zeniyacho_values_for_empty_column
                else:
                    next_df["12月"] = fukune_values_for_empty_column

            next_df.iloc[:, 1:] = next_df.iloc[:, 1:].apply(lambda x: x.str.replace(',', ''), axis=1)
            next_df.iloc[:, 1:] = next_df.iloc[:, 1:].apply(pd.to_numeric, errors='coerce')
            next_df["Row_Sum"] = next_df.iloc[:, 1:].sum(axis=1)
            next_df.fillna(0, inplace=True)
            if hotel_name == 'zeniyacho':
                column_13_sum_except_rows_1_and_9 = next_df.loc[(next_df.index != 0)
                                                                & (next_df.index != 8), 'Row_Sum'].sum()
                next_df.loc[8, "Row_Sum - 10%"] = next_df.iloc[8, 1:-1].sum() * 0.9
                next_total = column_13_sum_except_rows_1_and_9 + next_df.loc[8, "Row_Sum - 10%"]
                
                row_8_10 = []
                row_8 = next_df.iloc[8, 1:13]
                for i in row_8:
                    i = i * 0.1
                    i = int(i)
                    row_8_10.append(i)
                row_1 = next_df.iloc[0, 1:13]
                result = []
                for a, b in zip(row_1, row_8_10):
                    result.append(a - b)
                next_df.iloc[0, 1:13] = result
                
            else:
                column_13_sum_except_rows_1_and_9 = next_df.loc[(next_df.index != 0)
                                                                & (next_df.index != 7), 'Row_Sum'].sum()
                next_df.loc[7, "Row_Sum - 10%"] = next_df.iloc[7, 1:-1].sum() * 0.9
                next_total = column_13_sum_except_rows_1_and_9 + next_df.loc[7, "Row_Sum - 10%"]
                
                row_7_10 = []
                row_7 = next_df.iloc[7, 1:13]
                for i in row_7:
                    i = i * 0.1
                    i = int(i)
                    row_7_10.append(i)
                row_1 = next_df.iloc[0, 1:13]
                result = []
                for a, b in zip(row_1, row_7_10):
                    result.append(a - b)
                next_df.iloc[0, 1:13] = result

            next_column_15_sum = next_total
            sheet_name = f'{hotel_name} analysis {str(current_year)} {str(sheet_week)}'
            current_start_row = 3
            current_start_col = 0
            next_start_row = 23
            next_start_col = 0
            with pd.ExcelWriter(f'earnings_report_kyoto_machiyas/earnings_report_kyoto_machiyas_{hotel_name} analysis.xlsx', mode='a', if_sheet_exists='overlay') as writer:
                current_df.to_excel(writer, sheet_name=sheet_name, startrow=current_start_row,
                                    startcol=current_start_col, index=False)
                next_df.to_excel(writer, sheet_name=sheet_name, startrow=next_start_row, startcol=next_start_col,
                                 index=False)
                existing_sheets = writer.book.sheetnames

                existing_sheets.remove(sheet_name)
                existing_sheets.insert(0, sheet_name)

                writer.book._sheets = [writer.book[sheet_name] for sheet_name in existing_sheets]

    def individual_cell_values(hotel_name):

        workbook = xl.load_workbook(f'earnings_report_kyoto_machiyas/earnings_report_kyoto_machiyas_{hotel_name} analysis.xlsx')
        sheet = workbook[f'{hotel_name} analysis {str(current_year)} {str(current_week_of_year)}']
        previous_sheet = workbook[f'{hotel_name} analysis {str(current_year)} {str(previous_week_of_year)}']
        try:
            if workbook[f'{hotel_name} analysis {str(previous_year)} {str(52)}']:
                previous_sheet_from_previous_year = workbook[f'{hotel_name} analysis {str(previous_week_of_year)}']
        except KeyError:
            pass
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        grey_fill = PatternFill(start_color="2F2F2F", end_color="2F2F2F", fill_type="solid")
        green_fill = PatternFill(start_color="CDE0CD", end_color="CDE0CD", fill_type="solid")
        orange_fill = PatternFill(start_color="FFE3D4", end_color="FFE3D4", fill_type="solid")
        light_blue_fill = PatternFill(start_color="DAEDF4", end_color="DAEDF4", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))
        sheet['A1'] = f'Kyoto Machiya {hotel_name.capitalize()}'
        sheet['A1'].font = Font(size=22, bold=True)
        sheet['A2'] = f'{str(current_year)}'
        sheet['A2'].font = Font(size=22, bold=True)
        sheet['A22'] = f'{str(next_year)}'
        sheet['A22'].font = Font(size=22, bold=True)
        sheet['A5'] = "Total"
        sheet['A25'] = "Total"
        sheet['O18'] = f'{int(current_column_15_sum)}'
        sheet['N19'] = 'Direct Booking'
        sheet['O19'] = 'Please add amount'
        sheet['O38'] = f'{int(next_column_15_sum)}'
        sheet['N39'] = f'Direct Booking'
        sheet['O39'] = 'Please add amount'
        if current_week_of_year == 1:
            try:
                previous_sheet_from_previous_year = workbook[f'{hotel_name} analysis {previous_year} 52']
                previous_week_this_year_total = previous_sheet_from_previous_year['O18'].value or 0
                previous_week_next_year_total = previous_sheet_from_previous_year['O38'].value or 0
            except KeyError:
                previous_week_this_year_total = 0
                previous_week_next_year_total = 0
        else:
            try:
                workbook[f'{hotel_name} analysis {str(current_year)} {str(previous_week_of_year)}']
                previous_week_this_year_total = previous_sheet['O18'].value or 0
                previous_week_next_year_total = previous_sheet['O38'].value or 0
            except KeyError:
                previous_week_this_year_total = 0
                previous_week_next_year_total = 0
        sheet['O20'].fill = grey_fill
        sheet['O21'] = previous_week_this_year_total
        sheet['N21'] = 'Last Week'
        improved_week_this_year = int(current_column_15_sum) - int(previous_week_this_year_total)
        
        sheet['O22'] = improved_week_this_year
        sheet['O22'].fill = green_fill
        sheet['O22'].border = border
        
        sheet['N22'] = f'{current_year} Improved Week'
        sheet['O41'] = previous_week_next_year_total
        sheet['O40'].fill = grey_fill
        improved_week_next_year = int(next_column_15_sum) - int(previous_week_next_year_total)
        
        sheet['O42'] = improved_week_next_year
        sheet['O42'].fill = green_fill
        sheet['O42'].border = border
        this_year_total_improved_figure = int(current_column_15_sum) - int(previous_week_this_year_total)
        formatted_this_year_improved_total_figure = format(this_year_total_improved_figure, ',')
        next_year_total_improved_figure = int(next_column_15_sum) - int(previous_week_next_year_total)
        formatted_next_year_improved_total_figure = format(next_year_total_improved_figure, ',')
        total_improved_week_num = int(sheet['O22'].value) + int(sheet['O42'].value)
        formatted_total_improved_week_num = format(total_improved_week_num, ',')
        sheet['O42'] = formatted_next_year_improved_total_figure
        sheet['O22'] = formatted_this_year_improved_total_figure
        
        sheet['O45'] = formatted_total_improved_week_num
        sheet['O45'].fill = green_fill
        sheet['O45'].border = border
        
        sheet['N45'] = 'Total Improved Week'    
            
        sheet['N41'] = 'Last Week'
        sheet['N42'] = f'{next_year} Improved Week'
        
        sheet['N18'] = f'{current_year} Tentative Total'
        sheet['N18'].fill = highlight_fill
        
        sheet['O18'].fill = highlight_fill 
        
        sheet['N38'] = f'{next_year} Tentative Total'
        sheet['N38'].fill = highlight_fill
        
        sheet['O38'].fill = highlight_fill
        
        for column in range(1, 15):
            cell = sheet.cell(row=5, column=column)
            cell.fill = light_blue_fill
        for column in range(1, 15):
            cell = sheet.cell(row=25, column=column)
            cell.fill = light_blue_fill
            
        # The row for the orange line is different between fukune and zeniyacho
        if hotel_name == 'fukune':
            for column in range(1, 16):
                cell = sheet.cell(row=12, column=column)
                cell.fill = orange_fill
            for column in range(1, 16):
                cell = sheet.cell(row=32, column=column)
                cell.fill = orange_fill
        else:
            for column in range(1, 16):
                cell = sheet.cell(row=13, column=column)
                cell.fill = orange_fill
            for column in range(1, 16):
                cell = sheet.cell(row=33, column=column)
                cell.fill = orange_fill

        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        workbook.save(f'earnings_report_kyoto_machiyas/earnings_report_kyoto_machiyas_{hotel_name} analysis.xlsx')

    def delete_rightmost_sheet(file_path):
        wb = xl.load_workbook(file_path)
        if len(wb.sheetnames) >= 13:
            rightmost_sheet = wb.sheetnames[-1]
            wb.remove(wb[rightmost_sheet])
            wb.save(file_path)
            print(f"The rightmost sheet '{rightmost_sheet}' has been deleted.")
        else:
            print("There are less than 11 sheets in the Excel file. No sheets were deleted.")

    get_tables('zeniyacho', zeniyacho_contract_code, zeniyacho_account_name, zeniyacho_password,
               zeniyacho_neppan_site_url, current_week_of_year)
    individual_cell_values('zeniyacho')
    delete_rightmost_sheet('earnings_report_kyoto_machiyas/earnings_report_kyoto_machiyas_zeniyacho analysis.xlsx')
    get_tables('fukune', fukune_contract_code, fukune_account_name, fukune_password,
               fukune_neppan_site_url, current_week_of_year)
    individual_cell_values('fukune')
    delete_rightmost_sheet('earnings_report_kyoto_machiyas/earnings_report_kyoto_machiyas_fukune analysis.xlsx')

    email_sender = os.environ.get('KYOTO_MACHIYA_FUKUNE_GMAIL_ADDRESS')
    email_password = os.environ.get('KYOTO_MACHIYA_FUKUNE_GMAIL_ADDRESS_PASSWORD')
    email_receivers = [os.environ.get('BOSS_EMAIL'), os.environ.get('STAFF_EMAIL')]
    email_bcc = [os.environ.get('ALEX_EMAIL')]

    subject = "Kyoto Machiya Earnings Data"
    body = f"""
    <html>
        <body>
            <img src="https://kyotomachiyas.com/wp-content/uploads/2022/10/Logo-Transparent-Logo.png" 
            width="25%" height="25%" align="center">
            <br>
            <h1>Earnings Data</h1>
            <h2>week: {current_week_of_year}</h2>
        </body>
    </html>
    """
    
    data_files = ["earnings_report_kyoto_machiyas/earnings_report_kyoto_machiyas_fukune analysis.xlsx", "earnings_report_kyoto_machiyas/earnings_report_kyoto_machiyas_zeniyacho analysis.xlsx"]

    em = EmailMessage()
    em['From'] = email_sender
    em['To'] = email_receivers
    em['Subject'] = subject
    em.set_content(body, subtype='html')

    for data_file in data_files:
        with open(data_file, "rb") as f:
            em.add_attachment(f.read(), filename=data_file, maintype="application", subtype="octet-stream")
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
        smtp.login(email_sender, email_password)
        smtp.sendmail(email_sender, email_receivers + email_bcc, em.as_string())


if __name__ == '__main__':
    main()
